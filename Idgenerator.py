import pandas as pd
from imdb import IMDb
from openpyxl import load_workbook
import time

def get_imdb_id(movie_name):
    ia = IMDb()
    try:
        search_results = ia.search_movie(movie_name)
        if search_results:
            first_result = search_results[0]
            return "tt" + first_result.movieID
        else:
            return None
    except Exception as e:
        print(f"Error processing movie '{movie_name}': {e}")
        return None

# Function to add IMDb IDs to the existing Excel file
def add_imdb_ids_to_excel(file_path):
    # Load Excel file
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Iterate over rows and add IMDb IDs
    for row in range(2, ws.max_row + 1):
        movie_title = ws.cell(row=row, column=1).value
        imdb_id = get_imdb_id(movie_title)
        if imdb_id:
            ws.cell(row=row, column=2, value=imdb_id)
            print(f"Added: {movie_title} - IMDb ID: {imdb_id}")
        else:
            print(f"No IMDb ID found for: {movie_title}")
        
        # Rate limit to avoid hitting IMDb API limits
        time.sleep(1)  # 1-second delay between requests
    
    # Save the updated Excel file
    wb.save(file_path)

    print("IMDb IDs added to the existing Excel file.")

# Example usage
file_path = 'tt.xlsx'  # Modify this with the actual path to your Excel file
add_imdb_ids_to_excel(file_path)
